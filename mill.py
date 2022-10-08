
from tkinter import *
from tkinter import messagebox
from tkinter.messagebox import askyesno, askquestion, showinfo
import tksheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl
import os
import shutil
import datetime
import xlwings as xw
import time 
import xlrd

window = Tk()
window.title("Percubaan pertama") #label atas
window.geometry('800x500') #setting window size

selamat_datang = Label(window, text="Selamat datang!!", font=("Arial Bold", 20)) 
selamat_datang.grid(column=0, row=0, sticky=W) #sticky E=kekanan, W=kekiri

btn_report = Button(window, text="Teruskan Mill Production", command=window.destroy) 
btn_report.grid(column=0, row=2)

mainloop()

#path kedudukan templet file
def file_templet_path():
    global templet_file, current_folder, templet_folder
    current_folder = os.getcwd()

    d = 0 
    for i in os.listdir(current_folder):
        if i == 'templet':
            templet_folder = f'{current_folder}/{i}'
        else:
            d += 1

    d = 0
    for i in os.listdir(templet_folder):
        if i == 'templet.xlsx':
            templet_file = f'{templet_folder}/{i}'
        else:
            d += 1


#data dari templet excel
def load_file_templet():
    global wb_file_templet,ws_file_templet_summary, ws_file_templet_parent, ws_file_templet_pd, ws_file_templet_child, list_header_col_templet_parent, list_header_col_templet_child, list_mill_templet, column_last_templet_pd, dict_summary,ws_file_templet_front, ws_xw_file_templet_front,ws_xw_file_templet_pd,ws_xw_file_templet_summary, wb_xw_file_templet
    wb_file_templet = load_workbook(templet_file)
    wb_xw_file_templet = xw.Book(templet_file)
    # print(wb_file_templet.sheetnames)
    ws_file_templet_parent = wb_file_templet['Sheet4']
    ws_file_templet_child = wb_file_templet['Quality_K']
    ws_file_templet_pd = wb_file_templet['Sheet2']
    ws_file_templet_summary = wb_file_templet['Sheet1']
    ws_file_templet_front = wb_file_templet['Mills Daily Mgmt Report']
    ws_xw_file_templet_front = wb_xw_file_templet.sheets['Mills Daily Mgmt Report']
    ws_xw_file_templet_pd = wb_xw_file_templet.sheets['Sheet2']
    ws_xw_file_templet_summary = wb_xw_file_templet.sheets['Sheet1']

    a = 2
    list_header_col_templet_parent = []
    while ws_file_templet_parent[f'{gcl(a)}2'].value != None:
        header_column_parent = ws_file_templet_parent[f'{gcl(a)}2'].value
        list_header_col_templet_parent.append(header_column_parent)
        a += 1
    
    a = 2
    list_header_col_templet_child = []
    while ws_file_templet_child[f'{gcl(a)}2'].value != None:
        header_column_child = ws_file_templet_child[f'{gcl(a)}2'].value
        list_header_col_templet_child.append(header_column_child)
        a += 1

    
    # get data mill from excel
    row_mill = 5
    list_mill_templet = []
    while ws_file_templet_pd[f'E{str(row_mill)}'].value != None:
        row_mill += 1

    for i in range(5,row_mill):
        list_mill_templet.append(ws_file_templet_pd[f'E{str(i)}'].value)
    
    #last column pd dekat templet 
    column_last_templet_pd = 6 
    while ws_xw_file_templet_pd.range(f'{gcl(column_last_templet_pd)}6').value != 'P' and ws_xw_file_templet_pd.range(f'{gcl(column_last_templet_pd)}6').value != 'X':
        column_last_templet_pd += 1
    column_last_templet_pd = column_last_templet_pd - 6

    #get data from summary
    dict_summary = {}
    data_list_OER = ws_xw_file_templet_summary.range('B7:C7').value
    data_list_KER = ws_xw_file_templet_summary.range('B11:C11').value
    data_list_CPO = ws_xw_file_templet_summary.range('B15:E15').value
    data_list_PK = ws_xw_file_templet_summary.range('B19:E19').value

    dict_summary['OER'] = data_list_OER
    dict_summary['KER'] = data_list_KER
    dict_summary['CPO'] = data_list_CPO
    dict_summary['PK'] = data_list_PK


file_templet_path()
load_file_templet()
# print(list_header_col_templet_parent)
print(len(list_header_col_templet_parent))
# print(list_header_col_templet_child)

#dapatkan pd data
def pd():
    top = Tk()
    top.title('Mill Production')
    sheet = tksheet.Sheet(top, data=[[]])  # set up empty table
    sheet.headers(list_header_col_templet_parent)
    sheet.grid(column=0,row=0)
    sheet.set_sheet_data([["" for cj in range(2)] for ri in range(40)]) #cari the exact row and column
    sheet.enable_bindings(("single_select",
                            "row_select",
                            "column_width_resize",
                            "arrowkeys",
                            "right_click_popup_menu",
                            "rc_select",
                            "rc_insert_row",
                            "rc_delete_row",
                            "copy",
                            "cut",
                            "paste",
                            "delete",
                            "undo",
                            "edit_cell"
                            ))

    def data():
        btn_submit_bd_data = Button(sheet, text="Submit", command=hasil) #create button + color dlm kotak + color huruf
        btn_submit_bd_data.grid(column=1, row=1, sticky=W)

    def hasil():
        global data_parent, index_kosong
        data_parent = sheet.get_sheet_data(False, False, False)
        
        #nak buat try and excpet nanti 
        a = -1
        index_kosong = []
        for i in range(len(data_parent)):
            a += 1
            if data_parent[i][1] not in list_mill_templet:
                index_kosong.append(a)

        for i in reversed(index_kosong):
            data_parent.remove(data_parent[i])

        btn_submit_bd = Button(sheet, text="Submit", command=top.destroy) #create button + color dlm kotak + color huruf
        btn_submit_bd.grid(column=1, row=1, sticky=W)

    data()
    top.mainloop()
# print(data_parent)
# print(len(data_parent))


#dapatkan data pk
def pk():
    top = Tk()
    top.title('PK Quality')
    sheet = tksheet.Sheet(top, data=[[]])  # set up empty table
    sheet.headers(list_header_col_templet_child)
    sheet.grid(column=0,row=0)
    sheet.set_sheet_data([["0" for cj in range(2)] for ri in range(40)]) #cari the exact row and column
    sheet.enable_bindings(("single_select",
                            "row_select",
                            "column_width_resize",
                            "arrowkeys",
                            "right_click_popup_menu",
                            "rc_select",
                            "rc_insert_row",
                            "rc_delete_row",
                            "copy",
                            "cut",
                            "paste",
                            "delete",
                            "undo",
                            "edit_cell"
                            ))

    def data():
        btn_submit_bd_data = Button(sheet, text="Submit", command=hasil) #create button + color dlm kotak + color huruf
        btn_submit_bd_data.grid(column=1, row=1, sticky=W)

    def hasil():
        global data_child
        data_child = sheet.get_sheet_data(False, False, False)

        for i in reversed(index_kosong):
            data_child.remove(data_child[i])

        btn_submit_bd = Button(sheet, text="Submit", command=top.destroy) #create button + color dlm kotak + color huruf
        btn_submit_bd.grid(column=1, row=1, sticky=W)

    data()
    top.mainloop()
# print(data_parent)
# print(data_child)

pd()
pk()

#byk lagi function nak letak 

tarikh_parent = data_parent[0][0]
list_tarikh = tarikh_parent.split('/')
hari = list_tarikh[0] #dlm bentuk string
bulan = list_tarikh[1]
tahun = list_tarikh[2]
print(tarikh_parent, list_tarikh, hari, bulan, tahun)

if len(bulan) != 2:
    number = int(bulan)
    bulan = f"{number:02d}"

if len(hari) != 2:
    number = int(hari)
    hari = f"{number:02d}"

folder_hari_t = hari + bulan + tahun
#folder text bulan belum lagi

tarikh = datetime.datetime(int(tahun), int(bulan), int(hari))
text_bulan = tarikh.strftime("%b")

tarikh_s = tarikh - datetime.timedelta(days=1)
hari_s = tarikh_s.strftime("%d")
bulan_s = tarikh_s.strftime("%m")
tahun_s = tarikh_s.strftime("%Y")
text_bulan_s = tarikh_s.strftime("%b")

if len(bulan_s) != 2:
    number = int(bulan_s)
    bulan_s = f"{number:02d}"

if len(hari_s) != 2:
    number = int(hari_s)
    hari_s = f"{number:02d}"

folder_hari_st = hari_s + bulan_s + tahun_s

#check tarikh
for i in range(len(data_parent)):
    if tarikh_parent != data_parent[i][0]:
        # print('error', data_parent[i][1]) #popup message #ulang balik pd data
        messagebox.showerror('ERROR', 'Ada error dekat tarikh')
        continue

#listparentmill
list_mill_parent = []
for i in range(len(data_parent)):
    list_mill_parent.append(data_parent[i][1])
print(list_mill_parent)

#notcreate
list_not_create = []
for i in list_mill_templet:
    if i not in list_mill_parent:
        list_not_create.append(i)
print(list_not_create)

#notposted 
list_not_posted = []
for i in range(len(data_parent)):
    print(data_parent[i][1], data_parent[i][2])
    if data_parent[i][2] != 'Posted':
        list_not_posted.append(data_parent[i][1])
print(list_not_posted)
#success

#checknotposted dptkn index
index_not_posted = []
for i in range(len(list_not_posted)):
    index_not_posted.append(list_mill_parent.index(list_not_posted[i]))

index_not_create = []
for i in range(len(list_not_create)):
    index_not_create.append(list_mill_templet.index(list_not_create[i]))


#GUNA OS
# def copy_paste_file2():
#     for file_name in templet_folder:
#         if file_name == 'templet.xlsx':
#             file_source = f'{folder_hari_t}.xlsx'
#             folder_des = f'{folder_hari}/{file_source}'
#         else:
#             file_source =f'{file_name}'
#             folder_des = f'{folder_hari}/{file_source}'
        
#         if os.path.isfile(file_source):
#             shutil.copy(file_source, folder_des)
#             print('copied', file_source)
#     #settel copypaste

        

#tukar nama file
for i in os.listdir(templet_folder):
    templet_file_lain = f'{templet_folder}/{i}'
    split_folder = templet_file_lain.split('/')
    split_file = split_folder[-1].split(' ')
    print(split_file[0])
    if split_file[0] == '1.OER':
        split_folder[-1] = f'1.OER Achievement for {text_bulan} {tahun}-{folder_hari_t}.xlsx'
        file_oer = f'1.OER Achievement for {text_bulan} {tahun}-{folder_hari_t}.xlsx'
        oer_templet = '/'.join(split_folder)
        os.rename(templet_file_lain, oer_templet)
        print(oer_templet)
    elif split_file[0] == '2.KER':
        split_folder[-1] = f'2.KER Achievement for {text_bulan} {tahun}-{folder_hari_t}.xlsx'
        file_ker = f'2.KER Achievement for {text_bulan} {tahun}-{folder_hari_t}.xlsx'
        ker_templet = '/'.join(split_folder)
        os.rename(templet_file_lain, ker_templet)
        print(ker_templet)
    elif split_file[0] == '3.FFA':
        split_folder[-1] = f'3.FFA Average for {text_bulan} {tahun}-{folder_hari_t}.xlsx'
        file_ffa = f'3.FFA Average for {text_bulan} {tahun}-{folder_hari_t}.xlsx'
        ffa_templet = '/'.join(split_folder)
        os.rename(templet_file_lain, ffa_templet)
        print(ffa_templet)
    else:
        print(f'failed tukar tarikh {i}')



list_folder_asal = os.listdir(current_folder)
# if tahun_s not in list_folder_asal:
#     folder_templet = f'{current_folder}/templet'
#     #terus pergi ke templet_file
#     #kene buat function lain utk 
#     print(folder_templet)
#     # copy_paste_file(folder_templet)

if tahun not in list_folder_asal:
    folder_tahun = f'{current_folder}/{tahun}'
    os.makedirs(folder_tahun)
else:
    folder_tahun = f'{current_folder}/{tahun}'

#buat untuk folder bulan 
list_folder_tahun = os.listdir(folder_tahun)
if bulan not in list_folder_tahun:
    folder_bulan = f'{folder_tahun}/{bulan}'
    os.makedirs(folder_bulan)
else:
    folder_bulan = f'{folder_tahun}/{bulan}'

#buat untuk folder hari
list_folder_bulan = os.listdir(folder_bulan)
if hari not in list_folder_bulan:
    folder_hari = f'{folder_bulan}/{hari}'
    os.makedirs(folder_hari)
else:
    folder_hari = f'{folder_bulan}/{hari}'



#copy tank 
def copy_tank():
    global tank_cpo, last_row_templet
    #get data
    last_row_templet = 5
    while ws_file_templet_pd[f'E{str(last_row_templet)}'].value != None:
        last_row_templet += 1

    first_row_templet = 5
    tank_cpo = [[]]*(last_row_templet - first_row_templet)

    tank_cpo = ws_xw_file_templet_pd.range(f'AG{first_row_templet}:BA{last_row_templet}').value


def paste_tank():
    bil_tank = 7
    tank_1 = 'CPO Stor. Tank 1 (MT)'
    tank_2 = 'CPO Stor. Tank 2 (MT)'

    start = 1
    while ws_file_templet_parent[f'{gcl(start)}2'].value != tank_1:
        start += 1
    r = 1
    while ws_file_templet_parent[f'{gcl(r)}2'].value != tank_2:
        r += 1
    beza = r-start

    col_check2 = []
    for i in range(bil_tank*beza):
        if i%3 == 0:
            col_check2.append(i)
    col_check = []
    for i in range(bil_tank*beza):
        if i not in col_check2:
            col_check.append(i)

    for e in range(len(list_mill_templet)):
        if ws_file_templet_parent[f'C{e+3}'].value in list_mill_templet:
            g = list_mill_templet.index(ws_file_templet_parent[f'C{e+3}'].value)
            for f in range(len(tank_cpo)):
                if f in col_check:
                    if ws_file_templet_parent[f'{gcl(f+start)}{g+3}'].value == 0 or ws_file_templet_parent[f'{gcl(f+start)}{g+3}'].value == '0' :
                        ws_file_templet_parent[f'{gcl(f+start)}{g+3}'] = tank_cpo[e][f]

            for j in col_check2:
                if ws_file_templet_parent[f'{gcl(j+start)}{g+3}'].value == 0:
                    ws_file_templet_parent[f'{gcl(j+start+1)}{g+3}'] = 0
                    ws_file_templet_parent[f'{gcl(j+start+2)}{g+3}'] = 0

        else:
            print('salah3')

# copypastyte
    # first_row_templet = 5
    # for i in range(last_row_templet-5):
    #     list_mill_parent[i] = 0
    #     first_row_templet += 1

    # tambah_row = 0
    # for row in range(5,last_row_templet):
    #     for col in range(34,54):
    #         tank_cpo[tambah_row].append(ws_file_templet_pd[f'{gcl(col)}{row}'].value)
    #     tambah_row += 1

copy_tank()
list_mill_parent
def ejas_negatif():
    m = 0
    for q in data_parent:
        for s in range(len(q)):
            k = q[s]
            if k[-1] == '-':
                k = k.strip(k[-1])
                k = f'-{k}'
            data_parent[m][s] = k
        m += 1


def paste_parent():
    row_pd = 3
    for i in data_parent:
        col_pd = 2
        for k in i:
            ws_file_templet_parent[f'{gcl(col_pd)}{row_pd}'] = k
            col_pd += 1
        row_pd += 1
        col_pd = 2

def paste_child():
    row_kq = 3
    for i in data_child:
        col_kq = 2
        for k in i:
            ws_file_templet_child[f'{gcl(col_kq)}{row_kq}'] = k
            col_kq +=1
        row_kq += 1
        col_kq = 2
    pass

def remove_row_posted(x):
    index_remove = list_mill_parent.index(x)
    data_parent.remove(data_parent[index_remove])
    list_mill_parent.remove(list_mill_parent[index_remove])

def ambil_data_semalam():
    global copy_data_semalam
    # print(oer_file, ker_file, ffa_file)
    copy_data_semalam = {}
    for k in list_ambil_data_s:
        row = list_mill_templet.index(k)+5
        copy_data_semalam[k] = ws_xw_file_templet_pd.range(f'F{row}:BH37').value
        print(copy_data_semalam)
    for k in list_ambil_data_s:
        list_kosong = [0,3,6,9,13,16,18,21,24,49,52]
        for i in list_kosong:
            if len(list_ambil_data_s) > 1:
                copy_data_semalam[k][0][i] = 0
            else:
                copy_data_semalam[k][i] = 0

    print(copy_data_semalam)

def paste_data_semalam():
    d = 3
    while ws_file_templet_parent[f'C{d}'].value != None:
        d += 1

    mill = copy_data_semalam.keys()
    s = []
    for cell in ws_file_templet_parent['C']:
        s.append(cell.value)

    for i in mill:
        if i not in s:
            ws_file_templet_parent[f'C{d}'] = i
            print(d)
        e = 5
        for x in copy_data_semalam.get(i): #dict copy
            ws_file_templet_parent[f'{gcl(d)}{e}'] = x
            e += 1
        d += 1

def ejas_summary():
    ws_file_templet_summary['A23'] = dict_summary['OER'][0]
    ws_file_templet_summary['B23'] = dict_summary['OER'][1]
    ws_file_templet_summary['A27'] = dict_summary['KER'][0]
    ws_file_templet_summary['B27'] = dict_summary['KER'][1]
    ws_file_templet_summary['A31'] = dict_summary['CPO'][0]
    ws_file_templet_summary['B31'] = dict_summary['CPO'][1]
    ws_file_templet_summary['C31'] = dict_summary['CPO'][2]
    ws_file_templet_summary['D31'] = dict_summary['CPO'][3]
    ws_file_templet_summary['A35'] = dict_summary['PK'][0]
    ws_file_templet_summary['B35'] = dict_summary['PK'][1]
    ws_file_templet_summary['C35'] = dict_summary['PK'][2]
    ws_file_templet_summary['D35'] = dict_summary['PK'][3]

def message_box(x,y):
    global kalau_yes
    kalau_yes = askyesno(f'{x}', f'{y}')

#confirmkan
title = "CONFRIMATION"
list_ambil_data_s = []
if len(list_not_posted) >= 1:
    for i in list_not_posted:
        content_valid = f'IS THE DATA VALID for {i}?'
        message_box(title, content_valid)
        if kalau_yes == False:
            content_valid = f'{i} Ambil data semalam (yes) or kosongkan (no)'
            message_box(title, content_valid)
            if kalau_yes == True:
                list_ambil_data_s.append(i)
            else:
                remove_row_posted(i)

if len(list_not_create) >= 1:
    for i in list_not_create:
        content_create = f'Data {i} not create, ambil data semalam?'
        message_box(title, content_create)
        if kalau_yes == True:
            list_ambil_data_s.append(i)
#succes

def copy_paste_file_lain():
    global wb_file_oer, wb_file_ker, wb_file_ffa
    wb_file_oer = xw.Book(oer_templet)
    wb_file_ker = xw.Book(ker_templet)
    wb_file_ffa = xw.Book(ffa_templet)
    ws_file_oer = wb_file_oer.sheets[text_bulan]
    ws_file_ker = wb_file_ker.sheets[text_bulan]
    ws_file_ffa = wb_file_ffa.sheets[text_bulan]
    ws_file_oer_ytd = wb_file_oer.sheets[f'FY{tahun_s}']  # bet
    ws_file_ker_ytd = wb_file_ker.sheets[f'FY{tahun_s}']
    ws_file_ffa_ytd = wb_file_ffa.sheets[f'FY{tahun_s}']
    wb_file_oer.sheets[f'FY{tahun_s}'].name = f'FY{tahun}'
    wb_file_ker.sheets[f'FY{tahun_s}'].name = f'FY{tahun}'
    wb_file_ffa.sheets[f'FY{tahun_s}'].name = f'FY{tahun}'
    

    if hari == '01' and bulan == '01':
        for i in range(1,13):
            tarikh_tipu = datetime.datetime(int(2022), int(i), int(1))
            text_bulan_tipu = tarikh_tipu.strftime("%b")
            wb_file_oer.sheets[text_bulan_tipu].range('F9:AL52').value = None
            wb_file_ker.sheets[text_bulan_tipu].range('F8:AL52').value = None
            wb_file_ffa.sheets[text_bulan_tipu].range('F10:AL52').value = None
            wb_file_oer.sheets[text_bulan_tipu]
        ws_file_oer_ytd.range('F9:S52').value = None
        ws_file_ker_ytd.range('F9:S52').value = None
        ws_file_ffa_ytd.range('E10:P52').value = None

        for i in range(2,13):
            tarikh_tipu = datetime.datetime(int(2022), int(i), int(1))
            text_bulan_tipu = tarikh_tipu.strftime("%b")
            wb_file_oer.sheets[text_bulan_tipu].visible = False
            wb_file_ker.sheets[text_bulan_tipu].visible = False
            wb_file_ffa.sheets[text_bulan_tipu].visible = False
        
    if hari == '01':
        tarikh_tipu = datetime.datetime(int(tahun), int(bulan), int(hari))
        text_bulan_tipu = tarikh_tipu.strftime("%b")
        wb_file_oer.sheets[text_bulan_tipu].visible = True
        wb_file_ker.sheets[text_bulan_tipu].visible = True
        wb_file_ffa.sheets[text_bulan_tipu].visible = True


    hari_pertama = 6
    hari_file = hari_pertama + int(hari) - 1
    bulan_pertama = 6
    bulan_file = bulan_pertama + int(bulan) - 1

    copy_today_oer = ws_xw_file_templet_front.range('AG9:AG52').value
    copy_mtd_oer = ws_xw_file_templet_front.range('AH9:AH52').value
    copy_ytd_oer = ws_xw_file_templet_front.range('AI9:AI52').value

    copy_today_ker = ws_xw_file_templet_front.range('AJ9:AJ52').value
    copy_mtd_ker = ws_xw_file_templet_front.range('AK9:AK52').value
    copy_ytd_ker = ws_xw_file_templet_front.range('AL9:AL52').value

    copy_today_ffa = ws_xw_file_templet_front.range('X10:X50').value
    copy_mtd_ffa = ws_xw_file_templet_front.range('Y10:Y50').value

    max_col_oer = ws_file_oer.range('A5').end('right').column
    max_col_ker = ws_file_ker.range('A4').end('right').column
    max_col_ffa = ws_file_ffa.range('A5').end('right').column
    max_col_oer_ytd = ws_file_oer_ytd.range('A5').end('right').column 
    max_col_ker_ytd = ws_file_ker_ytd.range('A5').end('right').column


    row_copy = 0
    for i in copy_today_oer:
        ws_file_oer.range(f'{gcl(hari_file)}{row_copy+9}').value = i
        row_copy += 1
    row_copy = 0
    for i in copy_mtd_oer:
        ws_file_oer.range(f'{gcl(max_col_oer)}{row_copy+9}').value = i
        row_copy += 1
    row_copy = 0
    for i in copy_mtd_oer:
        ws_file_oer_ytd.range(f'{gcl(bulan_file)}{row_copy+9}').value = i
        row_copy += 1
    row_copy = 0
    for i in copy_ytd_oer:
        ws_file_oer_ytd.range(f'{gcl(max_col_oer_ytd)}{row_copy+9}').value = i
        row_copy += 1
    row_copy = 0
    for i in copy_today_ker:
        ws_file_ker.range(f'{gcl(hari_file)}{row_copy+8}').value = i
        row_copy += 1
    row_copy = 0
    for i in copy_mtd_ker:
        ws_file_ker.range(f'{gcl(max_col_ker)}{row_copy+8}').value = i
        row_copy += 1
    row_copy = 0
    for i in copy_mtd_ker:
        ws_file_ker_ytd.range(f'{gcl(bulan_file)}{row_copy+9}').value = i
        row_copy += 1
    row_copy = 0
    for i in copy_ytd_ker:
        ws_file_ker_ytd.range(f'{gcl(max_col_ker_ytd)}{row_copy+9}').value = i
        row_copy += 1
    row_copy = 0
    for i in copy_today_ffa:
        ws_file_ffa.range(f'{gcl(hari_file)}{row_copy+10}').value = i
        row_copy += 1
    row_copy = 0
    for i in copy_mtd_ffa:
        ws_file_ffa.range(f'{gcl(max_col_ffa)}{row_copy+10}').value = i
        row_copy += 1
    row_copy = 0
    for i in copy_mtd_ffa:
        ws_file_ffa_ytd.range(f'{gcl(bulan_file-1)}{row_copy+10}').value = i
        row_copy += 1

def padam_data_semalam():
    for col in range(2,61):
        for row in range(3,40):
            ws_file_templet_parent[f'{gcl(col)}{row}'].value = None

    for col in range(2,33):
        for row in range(3,40):
            ws_file_templet_child[f'{gcl(col)}{row}'].value = None
    # ws_file_templet_parent['B3:BH39'].value = None
    # ws_file_templet_child['B3:AF39'].value = None

ambil_data_semalam()

ejas_negatif()

padam_data_semalam()

paste_parent()

paste_data_semalam() 

paste_child()

paste_tank()

#tukar tarikh 
ws_file_templet_front['B3'] = tarikh_parent

if hari == '01':
    ejas_summary()

copy_paste_file_lain()

ws_file_templet_pd.sheet_state = 'hidden'
ws_file_templet_parent.sheet_state = 'hidden'
ws_file_templet_child.sheet_state = 'hidden'


wb_xw_file_templet.close()

#save file 
time.sleep(1)
wb_file_templet.save(templet_file)
wb_file_oer.save()
wb_file_ker.save()
wb_file_ffa.save()
wb_file_templet.save(f'{folder_hari}/{folder_hari_t}.xlsx') #folder baru 
wb_file_oer.close()
wb_file_ffa.close()
wb_file_ker.close()


wb_file_oer = load_workbook(oer_templet)
wb_file_ker = load_workbook(ker_templet)
wb_file_ffa = load_workbook(ffa_templet)
ws_file_oer_fy = wb_file_oer[f'FY{tahun}']
ws_file_ker_fy = wb_file_ker[f'FY{tahun}']
ws_file_ffa_fy = wb_file_ffa[f'FY{tahun}']
ws_file_oer_mtd = wb_file_oer[text_bulan]
ws_file_ker_mtd = wb_file_ker[text_bulan]
ws_file_ffa_mtd = wb_file_ffa[text_bulan]

if hari == '01' and bulan == '01':
    for i in range(1,13):
        tarikh_tipu = datetime.datetime(int(2022), int(i), int(1))
        text_bulan_tipu = tarikh_tipu.strftime("%b")
        max_col_oer = wb_file_oer[text_bulan_tipu].max_column
        max_col_ker = wb_file_ker[text_bulan_tipu].max_column
        max_col_ffa = wb_file_ffa[text_bulan_tipu].max_column
        wb_file_oer[text_bulan_tipu].column_dimensions.group('M',gcl(max_col_oer-1),hidden=True)
        wb_file_ker[text_bulan_tipu].column_dimensions.group('M',gcl(max_col_ker-1),hidden=True)
        wb_file_ffa[text_bulan_tipu].column_dimensions.group('M',gcl(max_col_ffa-1),hidden=True)
    ws_file_oer_fy.column_dimensions.group('G','Q',hidden=True)
    ws_file_ker_fy.column_dimensions.group('G','Q',hidden=True)
    ws_file_ffa_fy.column_dimensions.group('F','P',hidden=True)

i = int(hari) + 5
m = int(bulan) + 5
ws_file_oer_mtd.column_dimensions.group(gcl(i),hidden=False)
ws_file_ker_mtd.column_dimensions.group(gcl(i),hidden=False)
ws_file_ffa_mtd.column_dimensions.group(gcl(i),hidden=False)
ws_file_oer_fy.column_dimensions.group(gcl(m),hidden=False) 
ws_file_ker_fy.column_dimensions.group(gcl(m),hidden=False)
ws_file_ffa_fy.column_dimensions.group(gcl(m-1),hidden=False)

#active sheeet
wb_file_oer.active = wb_file_oer[text_bulan]
wb_file_ker.active = wb_file_ker[text_bulan]
wb_file_ffa.active = wb_file_ffa[text_bulan]

#active cell
if hari == '01':
    ws_file_oer_mtd.sheet_view.selection[0].activeCell = 'E'
    ws_file_ker_mtd.sheet_view.selection[0].activeCell = 'E'
    ws_file_ffa_mtd.sheet_view.selection[0].activeCell = 'E'
else:
    ws_file_oer_mtd.sheet_view.selection[0].activeCell = 'AL'
    ws_file_ker_mtd.sheet_view.selection[0].activeCell = 'AL'
    ws_file_ffa_mtd.sheet_view.selection[0].activeCell = 'AL'


wb_file_oer.save(oer_templet)
wb_file_ker.save(ker_templet)
wb_file_ffa.save(ffa_templet)
wb_file_oer.save(f'{folder_hari}/{file_oer}')
wb_file_ker.save(f'{folder_hari}/{file_ker}')
wb_file_ffa.save(f'{folder_hari}/{file_ffa}')


showinfo('SUCCESS','File ready')
#hide column semua bila awal tahun
# unhide column setiap masuk hari 
# buat cell active semua tempat  


#tukar tarikh dekat depan
#nak buat if tarikh masuk bulan baru sheet1 templet kena ejas


#cek balik mcm mana kalau ada kosong kat tgh2 dia jadi takde semua data bawah dia 